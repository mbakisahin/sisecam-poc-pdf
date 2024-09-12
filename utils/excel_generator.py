import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import textwrap


class ExcelReportGenerator:
    """
    A class to handle the creation of Excel reports with comparison data, including hyperlinks, alternating row colors,
    and specific column formatting for the 'E' column and others.
    """

    def __init__(self):
        # Define gray and white fills for alternating row colors
        self.gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12, name='Arial')
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))
        self.link_font = Font(color="0000FF", underline="single", name='Calibri', size=11)
        self.custom_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    def apply_header_style(self, worksheet, col_count):
        """
        Applies header styles to the first row of the worksheet, including background color, font style,
        and alignment. Borders are added for the header cells.
        
        :param worksheet: The Excel worksheet to format the header.
        :param col_count: The number of columns to apply the header style to.
        """
        for col in range(1, col_count + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = self.header_font
            cell.border = self.thin_border

    def format_data_columns(self, worksheet, max_row):
        """
        Formats the data columns by alternating the row colors between white and gray, adding borders and setting
        text alignment for cells in columns 1 to 5.
        
        :param worksheet: The Excel worksheet to format.
        :param max_row: The maximum row to apply formatting.
        """
        for col_idx in range(1, 6):
            fill_color = self.white_fill if col_idx % 2 == 0 else self.gray_fill
            for row_idx in range(2, max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = fill_color
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def adjust_first_four_column_width(self, worksheet):
        """
        Adjusts the width of the first four columns to ensure readability. This method is specifically designed
        to set static column widths for the first four columns.

        :param worksheet: The Excel worksheet to adjust column widths.
        """
        worksheet.column_dimensions['A'].width = 20  # Directorate
        worksheet.column_dimensions['B'].width = 20  # Keyword
        worksheet.column_dimensions['C'].width = 20  # Date
        worksheet.column_dimensions['D'].width = 20  # Source (Original Document link)
        worksheet.column_dimensions['E'].width = 20  # Source (Original Document link)

    def add_combined_neighbor_links_and_comparisons_with_comments(self, worksheet, neighbor_urls,
                                                                  individual_comparisons):
        """
        Adds a single cell containing the links to similar documents and their comparison texts. The links are
        listed vertically within a comment in the cell, and the cell itself is styled with a custom background color.
        
        :param worksheet: The Excel worksheet to add the combined neighbor links.
        :param neighbor_urls: A list of URLs for similar documents.
        :param individual_comparisons: A list of comparison texts associated with each neighbor URL.
        """
        combined_cell = worksheet.cell(row=2, column=6)  # This will be in a single cell (e.g., F2)
        combined_cell.value = "Similar Documents"
        combined_cell.font = self.link_font
        combined_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        combined_cell.border = self.thin_border
        combined_cell.fill = self.custom_fill  # Apply background color to the cell

        # Prepare the details for all similar documents, add a heading
        combined_comment_text = "Similar Documents:\n\n"
        for idx, (neighbor_url, comparison_text) in enumerate(zip(neighbor_urls, individual_comparisons), 1):
            combined_comment_text += f"\nSimilar Document {idx}: {neighbor_url}\nComparison: {comparison_text}\n\n"

        combined_comment = Comment(combined_comment_text, "Comparison")
        combined_cell.comment = combined_comment

        # Adjust the column width
        worksheet.column_dimensions[get_column_letter(6)].width = 30  # Adjust the width of the "Similar Documents" column

    def create_excel(self, metadata, file_name='comparison_report.xlsx'):
        """
        Creates an Excel report based on provided metadata, adds formatting for headers, and populates data such as
        source links and comparison text. Includes hyperlinks, alternating row colors, and custom cell comments.

        :param metadata: A dictionary containing the data to populate the report (e.g., keyword, date, URL, etc.).
        :param file_name: The name of the Excel file to save (default is 'comparison_report.xlsx').
        """
        data = {
            'Related Directorate': ['Environment'],
            'Keyword': [metadata.get('keyword', 'N/A')],
            'Date': [metadata.get('date', 'N/A')],
            'Source': [''],
            'Key Differences': ['...']
        }

        df = pd.DataFrame(data)
        neighbor_urls = metadata.get('neighbor_urls', [])
        individual_comparisons = metadata.get('individual_comparisons', [])

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=0)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Apply header formatting
            self.apply_header_style(worksheet, len(df.columns))

            # Add the "Source" header and make the Original Document link clickable
            worksheet.cell(row=1, column=4).value = "Source"
            link_cell = worksheet.cell(row=2, column=4)
            original_url = metadata.get('url', '#')
            link_cell.value = 'Original Document'
            link_cell.hyperlink = original_url
            link_cell.font = self.link_font
            link_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            link_cell.border = self.thin_border

            # Add a comment for "Key Differences"
            key_diff_cell = worksheet.cell(row=2, column=5)  # "Key Differences" cell
            key_diff_cell.value = "..."
            key_diff_cell.font = Font(name='Arial', size=11)
            key_diff_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            key_diff_cell.border = self.thin_border

            key_diff_comment = Comment(metadata.get('combined_comparison', 'N/A'), "Key Differences")
            key_diff_cell.comment = key_diff_comment

            # Format data columns (alternating colors and borders)
            self.format_data_columns(worksheet, worksheet.max_row)

            # Add similar document links and comparison texts, and include them as comments in the cell
            self.add_combined_neighbor_links_and_comparisons_with_comments(worksheet, neighbor_urls,
                                                                           individual_comparisons)

            # Adjust the width of the first four columns
            self.adjust_first_four_column_width(worksheet)

